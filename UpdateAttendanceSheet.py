import openpyxl
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import mysql.connector

mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password=""
)

def markAttendance(rollset, course, branch, year, subject):
    currentDate = datetime.now()
    day = currentDate.day
    month = currentDate.strftime("%B")
    course = course.lower()
    branch = branch.lower()
    year = str(year)
    branch = branch.replace(" ","")
    fileName = "C:/Users/ASUS/Desktop/Resources/" + course + "/" + branch + "/" + year + "/" + subject + ".xlsx"
    rollno = []

    wb = load_workbook(filename = "C:/Users/ASUS/Desktop/Resources/" + course + "/"+ branch + "/" + year + "/" + subject + ".xlsx")
    if not month in wb.sheetnames:
        mycursor = mydb.cursor()
        mycursor.execute("Use "+course)
        query = "Select rollno, name from "+branch+year
        mycursor.execute(query)
        result = mycursor.fetchall()
        namelist = []
        for r in result:
            rollno.append(int(r[0]))
            namelist.append(r[1])

        book = load_workbook(filename = "C:/Users/ASUS/Desktop/Resources/" + course + "/"+ branch + "/" + year + "/" + subject + ".xlsx")
        writer = pd.ExcelWriter(fileName, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df = pd.read_excel(fileName)
        df.insert(0, "Roll No", rollno, True)
        df.insert(1, "Name", namelist, True)
        df.insert(2, "Total", [0]*len(rollno), True)
        df.to_excel(writer, sheet_name=month, index=False)
        writer.save()
        writer.close()

    wb = openpyxl.load_workbook(filename="C:/Users/ASUS/Desktop/Resources/" + course + "/" + branch + "/" + year + "/" + subject + ".xlsx")
    sheet = wb[month]
    students = sheet.max_row

    attendance = [0] * (students - 1)
    for roll in rollset:
        attendance[int(roll) - 1] = 1

    book = openpyxl.load_workbook(filename="C:/Users/ASUS/Desktop/Resources/" + course + "/" + branch + "/" + year + "/" + subject + ".xlsx")
    writer = pd.ExcelWriter(fileName, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df = pd.read_excel(fileName, sheet_name=month)
    total = df.iloc[:, -1]

    for i in range(0, len(total)):
        total[i] = total[i] + attendance[i]

    df.insert(2, str(currentDate.date()), attendance, True)
    df.to_excel(writer, sheet_name=month, index=False)
    writer.save()
    writer.close()






